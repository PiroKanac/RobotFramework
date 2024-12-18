# Import Package

import openpyxl

# Load Workbook

vk = openpyxl.load_workbook("C:\\Users\\niri\\Documents\\Test.xlsx")

print(vk.sheetnames)
print("Active Sheet is "+ vk.active.title)

# Create Object f Any sheet which you want to work

sh=vk["Test"]
print(sh.title)

# Read value from cell
print(sh['A3'].value)
# Second approach
c1= sh.cell(3,1)
print(c1.value)

# Find Rows and Columns data
rows = sh.max_row
columns = sh.max_column

print("Total Rows are - " + str(rows))
print("Total Columns are - " + str(columns))

for i in range(1,rows+1):
    for j in range(1, columns+1):
        c=sh.cell(i,j)
        print(c.value)

#Second approach
for r in sh['A1':'C4']:
    for c in r:
        print(c.value)